"""Excel output generation for AWS Services Reporter.

Generates Excel workbooks with multiple formatted sheets containing regional service data.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from ..core.config import Config


def create_excel_output(
    config: Config,
    regions: Dict[str, Dict[str, Any]],
    region_services: Dict[str, List[str]],
    service_names: Optional[Dict[str, str]] = None,
    enhanced_services: Optional[Dict[str, Dict[str, Dict[str, Any]]]] = None,
    metadata: Optional[Dict[str, Any]] = None,
    quiet: bool = False,
) -> bool:
    """Generate Excel workbook with multiple formatted sheets.

    Args:
        config: Configuration object containing output settings
        regions: Dictionary mapping region codes to region details dictionaries
        region_services: Dictionary mapping region codes to service lists
        service_names: Dictionary mapping service codes to display names
        enhanced_services: Dictionary with enhanced service metadata per region
        metadata: Optional metadata about the data fetch operation
        quiet: Suppress progress output if True

    Returns:
        True if Excel file was successfully created, False if dependencies missing or error occurred

    Creates:
        Excel workbook (.xlsx) with multiple sheets:
        - Regional Services: List of regions and their services
        - Service Matrix: Services vs regions availability grid
        - Region Summary: Regions with comprehensive details and service counts
        - Service Summary: Services with regional coverage statistics
        - Statistics: Summary data and metadata

    Requires:
        pandas and openpyxl packages for Excel functionality
    """
    output_path = (
        Path(config.output_dir) / "excel" / f"{Path(config.regions_filename).stem}.xlsx"
    )
    logger = logging.getLogger(__name__)

    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows

        if not quiet:
            print(f"  📝 Creating Excel output...")

        # Prepare data for Excel sheets
        all_service_codes = sorted(
            set().union(*region_services.values()) if region_services else []
        )

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Sheet 1: Regional Services
        ws_regional = wb.create_sheet("Regional Services")
        regional_data = []
        for region_code in sorted(regions.keys()):
            region_name = regions[region_code]["name"]
            services = sorted(region_services.get(region_code, []))
            for service_code in services:
                service_name = (
                    service_names.get(service_code, service_code)
                    if service_names
                    else service_code
                )

                regional_data.append(
                    [region_code, region_name, service_code, service_name]
                )

        # Add headers and data
        headers = ["Region Code", "Region Name", "Service Code", "Service Name"]
        ws_regional.append(headers)
        for row in regional_data:
            ws_regional.append(row)

        # Format headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        for cell in ws_regional[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Sheet 2: Service Matrix
        ws_matrix = wb.create_sheet("Service Matrix")
        sorted_regions = sorted(regions.keys())

        # Create matrix header
        matrix_header = ["Service"] + sorted_regions
        ws_matrix.append(matrix_header)

        # Create matrix data with full service names
        for service_code in all_service_codes:
            # Use full service name if available, otherwise use service code
            service_display = (
                service_names.get(service_code, service_code)
                if service_names
                else service_code
            )
            row = [service_display]
            for region_code in sorted_regions:
                available = (
                    "✓" if service_code in region_services.get(region_code, []) else "✗"
                )
                row.append(available)
            ws_matrix.append(row)

        # Format matrix headers
        for cell in ws_matrix[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Format availability indicators
        green_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        red_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )

        for row in ws_matrix.iter_rows(min_row=2):
            for cell in row[1:]:  # Skip service name column
                if cell.value == "✓":
                    cell.fill = green_fill
                elif cell.value == "✗":
                    cell.fill = red_fill

        # Sheet 3: Region Summary
        ws_summary = wb.create_sheet("Region Summary")

        # Create region summary data with enhanced details including launch dates
        summary_data = []
        for region_code in sorted(regions.keys()):
            region_details = regions[region_code]
            region_name = region_details["name"]
            launch_date = region_details.get("launch_date", "Unknown")
            launch_date_source = region_details.get("launch_date_source", "Unknown")
            announcement_url = region_details.get("announcement_url", "")
            service_count = len(region_services.get(region_code, []))
            az_count = region_details.get("az_count", 0)

            summary_data.append(
                [
                    region_code,
                    region_name,
                    launch_date,
                    launch_date_source,
                    announcement_url,
                    az_count,
                    service_count,
                ]
            )

        # Add headers and data
        summary_headers = [
            "Region Code",
            "Region Name",
            "Launch Date",
            "Launch Date Source",
            "Announcement URL",
            "Availability Zones",
            "Service Count",
        ]
        ws_summary.append(summary_headers)
        for row in summary_data:
            ws_summary.append(row)

        # Format summary headers
        for cell in ws_summary[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Format numeric columns with number formatting
        for row_idx in range(2, len(summary_data) + 2):
            # AZ count column (5th column)
            az_cell = ws_summary.cell(row=row_idx, column=5)
            az_cell.alignment = Alignment(horizontal="right")
            # Service count column (6th column)
            service_cell = ws_summary.cell(row=row_idx, column=6)
            service_cell.alignment = Alignment(horizontal="right")

        # Sheet 4: Service Summary
        ws_service_summary = wb.create_sheet("Service Summary")

        # Get all unique services and calculate their coverage
        all_service_codes = sorted(
            set().union(*region_services.values()) if region_services else []
        )
        total_regions = len(regions)

        # Create service summary data
        service_summary_data = []
        for service_code in all_service_codes:
            service_name = (
                service_names.get(service_code, service_code)
                if service_names
                else service_code
            )

            # Count how many regions have this service
            available_regions = [
                region_code
                for region_code, services in region_services.items()
                if service_code in services
            ]
            region_count = len(available_regions)
            coverage_percent = (
                (region_count / total_regions * 100) if total_regions > 0 else 0
            )

            service_summary_data.append(
                [service_code, service_name, region_count, f"{coverage_percent:.1f}%"]
            )

        # Add headers and data
        service_summary_headers = [
            "Service Code",
            "Service Name",
            "Region Count",
            "Coverage %",
        ]
        ws_service_summary.append(service_summary_headers)
        for row in service_summary_data:
            ws_service_summary.append(row)

        # Format service summary headers
        for cell in ws_service_summary[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Format numeric columns with right alignment
        for row_idx in range(2, len(service_summary_data) + 2):
            # Region count column (3rd column)
            region_count_cell = ws_service_summary.cell(row=row_idx, column=3)
            region_count_cell.alignment = Alignment(horizontal="right")
            # Coverage % column (4th column)
            coverage_cell = ws_service_summary.cell(row=row_idx, column=4)
            coverage_cell.alignment = Alignment(horizontal="right")

        # Sheet 5: Statistics
        ws_stats = wb.create_sheet("Statistics")

        # Calculate statistics
        total_service_instances = sum(
            len(services) for services in region_services.values()
        )
        avg_services = total_service_instances / len(regions) if regions else 0

        # Most/least available services
        service_coverage = {}
        for service_code in all_service_codes:
            available_regions = [
                region
                for region, services in region_services.items()
                if service_code in services
            ]
            # Use full name if available for display
            service_display = (
                service_names.get(service_code, service_code)
                if service_names
                else service_code
            )
            service_coverage[service_display] = len(available_regions)

        most_available = (
            max(service_coverage.keys(), key=lambda x: service_coverage[x])
            if service_coverage
            else "N/A"
        )
        least_available = (
            min(service_coverage.keys(), key=lambda x: service_coverage[x])
            if service_coverage
            else "N/A"
        )

        # Add statistics data
        stats_data = [
            ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Generator", "AWS Services Reporter v1.3.0"],
            [""],
            ["Summary Statistics", ""],
            ["Total Regions", len(regions)],
            ["Total Services", len(all_service_codes)],
            ["Total Service Instances", total_service_instances],
            ["Average Services per Region", f"{avg_services:.1f}"],
            [
                "Most Available Service",
                f"{most_available} ({service_coverage.get(most_available, 0)} regions)",
            ],
            [
                "Least Available Service",
                f"{least_available} ({service_coverage.get(least_available, 0)} regions)",
            ],
        ]

        if metadata:
            stats_data.extend(
                [
                    [""],
                    ["Fetch Metadata", ""],
                    ["Fetch Duration (seconds)", metadata.get("fetch_duration", "N/A")],
                    ["AWS Profile", metadata.get("aws_profile", "default")],
                ]
            )

        for row_data in stats_data:
            ws_stats.append(row_data)

        # Format statistics sheet
        for row in ws_stats.iter_rows():
            if row[0].value in ["Summary Statistics", "Fetch Metadata"]:
                for cell in row:
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(
                        start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
                    )

        # Auto-adjust column widths
        for ws in [ws_regional, ws_matrix, ws_summary, ws_service_summary, ws_stats]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Save workbook
        wb.save(output_path)

        # Get file size for both display and logging
        file_size = output_path.stat().st_size

        if not quiet:
            print(
                f"    ✓ Created Excel output ({file_size:,} bytes, {len(wb.worksheets)} sheets)"
            )

        logger.info(f"Created Excel output: {output_path} ({file_size:,} bytes)")
        return True

    except ImportError as e:
        if not quiet:
            print(f"  ⚠️  Excel output requires pandas and openpyxl: {e}")
        logger.warning(f"Excel output dependencies missing: {e}")
        return False
    except Exception as e:
        if not quiet:
            print(f"  ❌ Failed to create Excel output: {e}")
        logger.error(f"Failed to create Excel output: {e}")
        return False
