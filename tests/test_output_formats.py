"""Tests for output format generation (CSV, JSON, Excel)."""

import csv
import json
import os
import sys
import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

# Import functions to test
sys.path.insert(0, str(Path(__file__).parent.parent))
from aws_services_reporter.core.config import Config
from aws_services_reporter.output.csv_output import (
    create_region_summary_csv,
    create_regions_services_csv,
    create_services_regions_matrix_csv,
)
from aws_services_reporter.output.excel_output import create_excel_output
from aws_services_reporter.output.json_output import create_json_output


class TestOutputFormats:
    """Test cases for output format generation."""

    def setup_method(self):
        """Set up test fixtures."""
        self.temp_dir = tempfile.mkdtemp()
        self.config = Config(
            output_dir=self.temp_dir,
            regions_filename="test_regions.csv",
            matrix_filename="test_matrix.csv",
        )

        # Sample test data (updated to match current data structure)
        self.test_regions = {
            "us-east-1": {
                "name": "US East (N. Virginia)",
                "launch_date": "2006-08-25",
                "partition": "aws",
                "az_count": 3,
            },
            "eu-west-1": {
                "name": "Europe (Ireland)",
                "launch_date": "2007-10-10",
                "partition": "aws",
                "az_count": 3,
            },
            "ap-southeast-1": {
                "name": "Asia Pacific (Singapore)",
                "launch_date": "2010-04-28",
                "partition": "aws",
                "az_count": 3,
            },
        }

        self.test_region_services = {
            "us-east-1": ["ec2", "s3", "lambda"],
            "eu-west-1": ["ec2", "s3"],
            "ap-southeast-1": ["ec2"],
        }

        self.test_metadata = {"fetch_duration": 89.5, "aws_profile": "test-profile"}

    def teardown_method(self):
        """Clean up test fixtures."""
        # Remove all files and directories in temp directory
        import shutil

        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_create_regions_services_csv(self):
        """Test CSV region services output generation."""
        create_regions_services_csv(
            self.config, self.test_regions, self.test_region_services, quiet=True
        )

        csv_file = Path(self.temp_dir) / "csv" / "test_regions.csv"
        assert csv_file.exists()

        # Read and verify CSV content
        with open(csv_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        # Should have 6 rows (3 + 2 + 1 services)
        assert len(rows) == 6

        # Verify header (updated to match current CSV format)
        assert reader.fieldnames == [
            "Region Code",
            "Region Name",
            "Service Code",
            "Service Name",
        ]

        # Verify some specific entries
        us_east_rows = [row for row in rows if row["Region Code"] == "us-east-1"]
        assert len(us_east_rows) == 3

        services = [row["Service Code"] for row in us_east_rows]
        assert "ec2" in services
        assert "s3" in services
        assert "lambda" in services

    def test_create_services_regions_matrix_csv(self):
        """Test CSV services-regions matrix output generation."""
        create_services_regions_matrix_csv(
            self.config, self.test_regions, self.test_region_services, quiet=True
        )

        csv_file = Path(self.temp_dir) / "csv" / "test_matrix.csv"
        assert csv_file.exists()

        # Read and verify CSV content
        with open(csv_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        # Should have 3 service rows
        assert len(rows) == 3

        # Verify header includes all regions
        expected_headers = ["Service", "ap-southeast-1", "eu-west-1", "us-east-1"]
        assert reader.fieldnames == expected_headers

        # Create lookup for easy testing
        service_matrix = {row["Service"]: row for row in rows}

        # Test ec2 availability (all regions)
        assert service_matrix["ec2"]["us-east-1"] == "1"
        assert service_matrix["ec2"]["eu-west-1"] == "1"
        assert service_matrix["ec2"]["ap-southeast-1"] == "1"

        # Test s3 availability (2 regions)
        assert service_matrix["s3"]["us-east-1"] == "1"
        assert service_matrix["s3"]["eu-west-1"] == "1"
        assert service_matrix["s3"]["ap-southeast-1"] == "0"

        # Test lambda availability (1 region)
        assert service_matrix["lambda"]["us-east-1"] == "1"
        assert service_matrix["lambda"]["eu-west-1"] == "0"
        assert service_matrix["lambda"]["ap-southeast-1"] == "0"

    def test_create_region_summary_csv(self):
        """Test CSV region summary output generation."""
        create_region_summary_csv(
            self.config, self.test_regions, self.test_region_services, quiet=True
        )

        csv_file = Path(self.temp_dir) / "csv" / "region_summary.csv"
        assert csv_file.exists()

        # Read and verify CSV content
        with open(csv_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        # Should have 3 rows (one per region)
        assert len(rows) == 3

        # Verify header (updated to match current CSV format with AZ column)
        assert reader.fieldnames == [
            "Region Code",
            "Region Name",
            "Availability Zones",
            "Service Count",
        ]

        # Create lookup for easy testing
        region_summary = {row["Region Code"]: row for row in rows}

        # Verify specific entries
        assert region_summary["us-east-1"]["Region Name"] == "US East (N. Virginia)"
        assert region_summary["us-east-1"]["Service Count"] == "3"
        assert region_summary["us-east-1"]["Availability Zones"] == "3"

        assert region_summary["eu-west-1"]["Region Name"] == "Europe (Ireland)"
        assert region_summary["eu-west-1"]["Service Count"] == "2"

        assert (
            region_summary["ap-southeast-1"]["Region Name"]
            == "Asia Pacific (Singapore)"
        )
        assert region_summary["ap-southeast-1"]["Service Count"] == "1"

    def test_create_json_output(self):
        """Test JSON output generation with metadata."""
        result = create_json_output(
            self.config,
            self.test_regions,
            self.test_region_services,
            service_names=None,
            enhanced_services=None,
            metadata=self.test_metadata,
            quiet=True,
        )

        assert result is True

        json_file = Path(self.temp_dir) / "json" / "test_regions.json"
        assert json_file.exists()

        # Read and verify JSON content
        with open(json_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        # Verify structure
        assert "generated_at" in data
        assert "generator" in data
        assert "summary" in data
        assert "regions" in data
        assert "services" in data
        assert "metadata" in data

        # Verify generator info
        assert data["generator"]["name"] == "AWS Services Reporter"
        assert data["generator"]["version"] == "1.3.0"

        # Verify summary statistics
        summary = data["summary"]
        assert summary["total_regions"] == 3
        assert summary["total_services"] == 3
        assert summary["avg_services_per_region"] == 2.0  # (3+2+1)/3
        assert summary["most_available_service"] == "ec2"
        assert summary["least_available_service"] in [
            "lambda"
        ]  # Could be lambda (1 region)

        # Verify regions data
        regions_data = data["regions"]
        assert len(regions_data) == 3
        assert regions_data["us-east-1"]["name"] == "US East (N. Virginia)"
        assert regions_data["us-east-1"]["service_count"] == 3
        # Services are now a list of dictionaries with 'code' and 'name' keys
        service_codes = {
            service["code"] for service in regions_data["us-east-1"]["services"]
        }
        assert service_codes == {"ec2", "lambda", "s3"}

        # Verify services data with coverage
        services_data = data["services"]
        assert len(services_data) == 3

        assert services_data["ec2"]["available_in"] == 3
        assert services_data["ec2"]["coverage_percentage"] == 100.0

        assert services_data["s3"]["available_in"] == 2
        assert services_data["s3"]["coverage_percentage"] == 66.7

        assert services_data["lambda"]["available_in"] == 1
        assert services_data["lambda"]["coverage_percentage"] == 33.3

    def test_create_excel_output_success(self):
        """Test Excel output generation (when dependencies available)."""
        # Test the real Excel functionality if dependencies are available
        try:
            result = create_excel_output(
                self.config,
                self.test_regions,
                self.test_region_services,
                service_names=None,
                enhanced_services=None,
                metadata=self.test_metadata,
                quiet=True,
            )

            # If pandas/openpyxl are available, should succeed
            if result:
                excel_file = Path(self.temp_dir) / "excel" / "test_regions.xlsx"
                assert excel_file.exists()
                assert result is True

                # Test that Excel file has expected sheets
                try:
                    import openpyxl

                    wb = openpyxl.load_workbook(excel_file)
                    sheet_names = wb.sheetnames
                    expected_sheets = [
                        "Regional Services",
                        "Service Matrix",
                        "Region Summary",
                        "Service Summary",
                        "Statistics",
                    ]
                    assert len(sheet_names) == 5
                    for expected_sheet in expected_sheets:
                        assert expected_sheet in sheet_names

                    # Test region summary sheet content
                    ws_summary = wb["Region Summary"]
                    assert ws_summary["A1"].value == "Region Code"
                    assert ws_summary["B1"].value == "Region Name"
                    assert ws_summary["C1"].value == "Availability Zones"
                    assert ws_summary["D1"].value == "Service Count"

                    # Check data rows (should have 3 regions)
                    assert ws_summary.max_row == 4  # 1 header + 3 data rows
                    # Note: Excel might have extra columns, so just check we have at least 4
                    assert (
                        ws_summary.max_column >= 4
                    )  # At least: Region Code, Name, AZs, Service Count
                    wb.close()
                except ImportError:
                    # openpyxl not available for testing, but file was created
                    pass
            else:
                # Dependencies not available, test handled gracefully
                assert result is False
        except ImportError:
            # Dependencies not available, test handled gracefully
            result = create_excel_output(
                self.config,
                self.test_regions,
                self.test_region_services,
                self.test_metadata,
                quiet=True,
            )
            assert result is False

    def test_create_excel_output_missing_dependencies(self):
        """Test Excel output when dependencies are missing."""
        # Test when pandas import fails
        with patch(
            "builtins.__import__",
            side_effect=lambda name, *args: (
                ImportError("pandas not available")
                if name == "pandas"
                else __import__(name, *args)
            ),
        ):
            result = create_excel_output(
                self.config, self.test_regions, self.test_region_services, quiet=True
            )

            assert result is False

    def test_empty_data_handling(self):
        """Test output generation with empty data."""
        empty_regions = {}
        empty_services = {}

        # JSON should handle empty data gracefully
        result = create_json_output(
            self.config, empty_regions, empty_services, quiet=True
        )
        assert result is True

        json_file = Path(self.temp_dir) / "json" / "test_regions.json"
        with open(json_file, "r") as f:
            data = json.load(f)

        assert data["summary"]["total_regions"] == 0
        assert data["summary"]["total_services"] == 0
        assert data["summary"]["most_available_service"] is None

    def test_file_creation_errors(self):
        """Test handling of file creation errors."""
        # Try to write to non-existent directory in temp area
        import tempfile

        temp_parent = tempfile.mkdtemp()
        bad_config = Config(
            output_dir=f"{temp_parent}/non/existent/directory",
            regions_filename="test.csv",
        )

        # Should create directory and succeed
        result = create_json_output(
            bad_config, self.test_regions, self.test_region_services, quiet=True
        )
        assert result is True

        # Verify directory was created
        assert Path(f"{temp_parent}/non/existent/directory").exists()

        # Clean up
        import shutil

        shutil.rmtree(temp_parent)

    def test_large_data_handling(self):
        """Test output generation with large datasets."""
        # Create large test dataset (updated to match current data structure)
        large_regions = {
            f"region-{i:03d}": {
                "name": f"Region {i}",
                "launch_date": "2010-01-01",
                "partition": "aws",
                "az_count": i % 5 + 1,
            }
            for i in range(100)
        }
        large_services = {
            f"region-{i:03d}": [f"service-{j}" for j in range(i % 10)]
            for i in range(100)
        }

        # Test JSON output with large data
        result = create_json_output(
            self.config, large_regions, large_services, quiet=True
        )
        assert result is True

        # Verify file was created and has reasonable size
        json_file = Path(self.temp_dir) / "json" / "test_regions.json"
        assert json_file.exists()
        assert json_file.stat().st_size > 1000  # Should be reasonably large

    def test_special_characters_handling(self):
        """Test handling of special characters in region/service names."""
        special_regions = {
            "us-east-1": {
                "name": "US East (N. Virginia) - Special & Characters",
                "launch_date": "2006-08-25",
                "partition": "aws",
                "az_count": 3,
            },
            "eu-west-1": {
                "name": "Europe (Ireland) - Åccénted Chäracters",
                "launch_date": "2007-10-10",
                "partition": "aws",
                "az_count": 3,
            },
            "ap-se-1": {
                "name": "Asia Pacific - 中文/日本語",
                "launch_date": "2010-04-28",
                "partition": "aws",
                "az_count": 2,
            },
        }

        special_services = {
            "us-east-1": ["service-with-dashes", "service_with_underscores"],
            "eu-west-1": ["service.with.dots"],
            "ap-se-1": ["service@with@symbols"],
        }

        # Should handle special characters gracefully
        result = create_json_output(
            self.config, special_regions, special_services, quiet=True
        )
        assert result is True

        # Verify content is readable
        json_file = Path(self.temp_dir) / "json" / "test_regions.json"
        with open(json_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        assert "us-east-1" in data["regions"]
        # Services are now a list of dictionaries with 'code' and 'name' keys
        service_codes = {
            service["code"] for service in data["regions"]["us-east-1"]["services"]
        }
        assert "service-with-dashes" in service_codes


if __name__ == "__main__":
    pytest.main([__file__])
